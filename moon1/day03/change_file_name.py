import os
import openpyxl

files_list = os.listdir(r'D:\PycharmProjects\MyTest\moon1\day03\video')

wb = openpyxl.load_workbook(r'D:\PycharmProjects\MyTest\moon1\day03\rename.xlsx')
sh = wb['Sheet1']
key_list=[]
excel_list = []
for i in range(1,3):
    key_list.append(sh.cell(row=1,column=i).value)
for i in range(2,7):
    file_dict = {}
    for j in range(1,3):
        file_dict.update({key_list[j-1]:sh.cell(row=i,column=j).value})
    excel_list.append(file_dict)

for i in excel_list:
    for j in files_list:
        if j.split('.')[0] == i['old_name']:
            os.rename(os.path.join(r'D:\PycharmProjects\MyTest\moon1\day03\video',j),os.path.join(r'D:\PycharmProjects\MyTest\moon1\day03\video',i["new_name"]+'.pcm'))

print(os.path.join(r'D:\PycharmProjects\MyTest\moon1\day03\video','mm.txt'))


















































